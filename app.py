# Copyright 2018 Google LLC
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

# [START gae_python37_app]
import os
import io
import functools
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
import base64
import json
import traceback

from flask import Flask, request, g, session, redirect, url_for, render_template
from flask import render_template_string, jsonify, Response
from flask_github import GitHub

from sqlalchemy import create_engine, Column, Integer, String
from sqlalchemy.orm import scoped_session, sessionmaker
from sqlalchemy.ext.declarative import declarative_base

from datetime import datetime

# setup sqlalchemy

from config import *

engine = create_engine(DATABASE_URI)
db_session = scoped_session(sessionmaker(autocommit=False,
                                         autoflush=False,
                                         bind=engine))
Base = declarative_base()
Base.query = db_session.query_property()

class User(Base):
    __tablename__ = 'users'

    id = Column(Integer, primary_key=True)
    github_access_token = Column(String(255))
    github_id = Column(Integer)
    github_login = Column(String(255))

    def __init__(self, github_access_token):
        self.github_access_token = github_access_token

def init_db():
    Base.metadata.create_all(bind=engine)


# Create an app instance
class FlaskApp(Flask):
    def __init__(self, *args, **kwargs):
        super(FlaskApp, self).__init__(*args, **kwargs)
        self._activate_background_job()

    def _activate_background_job(self):
        init_db()


# If `entrypoint` is not defined in app.yaml, App Engine will look for an app
# called `app` in `main.py`.
app = FlaskApp(__name__)

app.config.from_object('config')

github = GitHub(app)


def verify_logged_in(fn):
    """
    Decorator used to make sure that the user is logged in
    """

    @functools.wraps(fn)
    def wrapped(*args, **kwargs):
        # If the user is not logged in, then redirect him to the "logged out" page:
        if not g.user:
            return redirect(url_for("loggedout"))
        return fn(*args, **kwargs)

    return wrapped

# Pages:


@app.before_request
def before_request():
    g.user = None
    if 'user_id' in session:
        g.user = User.query.get(session['user_id'])


@app.after_request
def after_request(response):
    db_session.remove()
    return response


@github.access_token_getter
def token_getter():
    user = g.user
    if user is not None:
        return user.github_access_token


@app.route('/github-callback')
@github.authorized_handler
def authorized(access_token):
    next_url = request.args.get('next') or url_for('home')
    if access_token is None:
        return redirect(next_url)

    user = User.query.filter_by(github_access_token=access_token).first()
    if user is None:
        user = User(access_token)
        db_session.add(user)

    user.github_access_token = access_token

    # Not necessary to get these details here
    # but it helps humans to identify users easily.
    g.user = user
    github_user = github.get('/user')
    user.github_id = github_user['id']
    user.github_login = github_user['login']

    db_session.commit()

    session['user_id'] = user.id
    return redirect(next_url)


@app.route('/login')
def login():
    if session.get('user_id', None) is None:
        return github.authorize(scope="user,repo")
    else:
        return 'Already logged in'


@app.route('/logout')
def logout():
    session.pop('user_id', None)
    return redirect(url_for('loggedout'))


@app.route("/loggedout")
def loggedout():
    """
    Displays the page to be shown to logged out users.
    """
    return render_template("loggedout.html")


@app.route('/user')
@verify_logged_in
def user():
    return jsonify(github.get('/user'))


# Pages for the app
@app.route('/')
@app.route('/home')
@verify_logged_in
def home():
    repositories = app.config['REPOSITORIES']
    return render_template('index.html',
                           login=g.user.github_login,
                           repos=repositories)


@app.route('/repo/<repo_key>')
@app.route('/repo/<repo_key>/<folder_path>')
@verify_logged_in
def repo(repo_key, folder_path=""):
    repositories = app.config['REPOSITORIES']
    repo_detail = repositories[repo_key]
    directories = github.get(
        f'repos/{repo_detail}/contents/{folder_path}'
    )
    #print(directories)
    dirs = []
    spreadsheets = []
    for directory in directories:
        if directory['type']=='dir':
            dirs.append(directory['name'])
        elif directory['type']=='file' and '.xlsx' in directory['name']:
            spreadsheets.append(directory['name'])
    if g.user.github_login in USERS_METADATA:
        user_initials = USERS_METADATA[g.user.github_login]
    else:
        print(f"The user {g.user.github_login} has no known metadata")
        user_initials = g.user.github_login[0:2]

    return render_template('repo.html',
                            login=g.user.github_login,
                            user_initials=user_initials,
                            repo_name = repo_key,
                            folder_path = folder_path,
                            directories = dirs,
                            spreadsheets = spreadsheets,
                            )


@app.route('/edit/<repo_key>/<folder>/<spreadsheet>')
@verify_logged_in
def edit(repo_key, folder, spreadsheet):
    repositories = app.config['REPOSITORIES']
    repo_detail = repositories[repo_key]
    spreadsheet_file = github.get(
        f'repos/{repo_detail}/contents/{folder}/{spreadsheet}'
    )
    base64_bytes = spreadsheet_file['content'].encode('utf-8')
    decoded_data = base64.decodebytes(base64_bytes)
    wb = openpyxl.load_workbook(io.BytesIO(decoded_data))
    sheet = wb.active

    header = [cell.value for cell in sheet[1] if cell.value]
    rows = []
    for row in sheet[2:sheet.max_row]:
        values = {}
        for key, cell in zip(header, row):
            values[key] = cell.value
        if any(values.values()):
            rows.append(values)
    if g.user.github_login in USERS_METADATA:
        user_initials = USERS_METADATA[g.user.github_login]
    else:
        print(f"The user {g.user.github_login} has no known metadata")
        user_initials = g.user.github_login[0:2]

    return render_template('edit.html',
                            login=g.user.github_login,
                            user_initials=user_initials,
                            all_initials=ALL_USERS_INITIALS,
                            repo_name = repo_key,
                            folder = folder,
                            spreadsheet_name=spreadsheet,
                            header=json.dumps(header),
                            rows=json.dumps(rows)
                            )


@app.route('/save', methods=['POST'])
@verify_logged_in
def save():
    repo_key = request.form.get("repo_key")
    folder = request.form.get("folder")
    spreadsheet = request.form.get("spreadsheet")
    row_data = request.form.get("rowData")
    commit_msg = request.form.get("commit_msg")

    repositories = app.config['REPOSITORIES']
    repo_detail = repositories[repo_key]

    try:
        row_data_parsed = json.loads(row_data)
        # Get the data, skip the first 'id' column
        first_row = row_data_parsed[0]
        header = [k for k in first_row.keys()]
        del header[0]
        del row_data_parsed[0] # Remove header row
        # Sort based on label
        row_data_parsed = sorted(row_data_parsed, key=lambda k: k['Label'] if k['Label'] else "")

        #print(row_data_parsed)

        wb = openpyxl.Workbook()
        sheet = wb.active
        #sheet.title="Definitions" # This creates a new sheet, no good

        for c in range(len(header)):
            sheet.cell(row=1, column=c+1).value=header[c]
            sheet.cell(row=1, column=c+1).font = Font(size=12,bold=True)
        for r in range(len(row_data_parsed)):
            row = [v for v in row_data_parsed[r].values()]
            del row[0]
            for c in range(len(header)):
                sheet.cell(row=r+2, column=c+1).value=row[c]
                # Set row background colours according to 'Curation status'
                # These should be kept in sync with those used in edit screen
                # TODO add to config maybe?
                if row[header.index("Curation status")]=="Discussed":
                    sheet.cell(row=r+2, column=c+1).fill = PatternFill(fgColor="FFE4B5", fill_type = "solid")
                elif row[header.index("Curation status")]=="In Discussion":
                    sheet.cell(row=r+2, column=c+1).fill = PatternFill(fgColor="fffacd", fill_type = "solid")
                elif row[header.index("Curation status")]=="To Be Discussed":
                    sheet.cell(row=r+2, column=c+1).fill = PatternFill(fgColor="eee8aa", fill_type = "solid")
                elif row[header.index("Curation status")]=="Ready":
                    sheet.cell(row=r+2, column=c+1).fill = PatternFill(fgColor="98fb98", fill_type = "solid")

        # Create version for saving
        spreadsheet_stream = io.BytesIO()
        wb.save(spreadsheet_stream)

        #base64_bytes = base64.b64encode(sample_string_bytes)
        base64_bytes = base64.b64encode(spreadsheet_stream.getvalue())
        base64_string = base64_bytes.decode("ascii")

        # Create a new branch to commit the change to (in case of simultaneous updates)
        response = github.get(f"repos/{repo_detail}/git/ref/heads/master")
        if not response or "object" not in response or "sha" not in response["object"]:
            raise Exception(f"Unable to get SHA for HEAD of master in {repo_detail}")
        sha = response["object"]["sha"]
        branch = f"{g.user.github_login}_{datetime.utcnow().strftime('%Y-%m-%d_%H%M')}"
        print("About to try to create branch in ",f"repos/{repo_detail}/git/refs")
        response = github.post(
            f"repos/{repo_detail}/git/refs", data={"ref": f"refs/heads/{branch}", "sha": sha},
            )
        if not response:
            raise Exception(f"Unable to create new branch {branch} in {repo_detail}")

        print("About to get SHA for the spreadsheet file",f"repos/{repo_detail}/contents/{folder}/{spreadsheet}")
        # Get the sha for the file
        response = github.get(f"repos/{repo_detail}/contents/{folder}/{spreadsheet}")
        if not response or "sha" not in response:
            raise Exception(
                f"Unable to get the current SHA value for {spreadsheet} in {repo_detail}/{folder}"
                )
        file_sha = response["sha"]

        # Commit changes to branch (replace code with sheet)
        data = {
            "message": commit_msg,
            "content": base64_string,
            "branch": branch,
        }
        if file_sha:
            data["sha"] = file_sha
        print("About to commit file to branch",f"repos/{repo_detail}/contents/{folder}/{spreadsheet}")
        response = github.put(f"repos/{repo_detail}/contents/{folder}/{spreadsheet}", data=data)
        if not response:
            raise Exception(
                f"Unable to commit addition of {spreadsheet} to branch {branch} in {repo_detail}"
            )

        # Create a PR for the change
        print("About to create PR from branch",)
        response = github.post(
            f"repos/{repo_detail}/pulls",
            data={
                "title": commit_msg,
                "head": branch,
                "base": "master",
            },
        )
        if not response:
            raise Exception(f"Unable to create PR for branch {branch} in {repo_detail}")

        # Merge the created PR
        print("About to merge created PR")
        response = github.post(
            f"repos/{repo_detail}/merges",
            data={
                "head": branch,
                "base": "master",
                "commit_message": commit_msg
            },
        )
        if not response:
            raise Exception(f"Unable to merge PR from branch {branch} in {repo_detail}")


        # Delete the branch again
        print ("About to delete branch",f"repos/{repo_detail}/git/refs/heads/{branch}")
        response = github.delete(
            f"repos/{repo_detail}/git/refs/heads/{branch}")
        if not response:
            raise Exception(f"Unable to delete branch {branch} in {repo_detail}")

        print ("Save succeeded")
        # TODO Figure out responses and send message. Options are Success / Save failure / Merge failure
        return ( "Success", 200 )

    except Exception as err:
        print(err)
        traceback.print_exc()
        return (
            format(err),
            400,
        )




if __name__ == "__main__":        # on running python app.py
    app.run(debug=app.config["DEBUG"], port=8080)        # run the flask app

# [END gae_python37_app]
